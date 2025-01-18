import csv
import io
from pprint import pprint
from zipfile import ZipFile
from openpyxl import load_workbook
from pypdf import PdfReader


# Читаем csv-файл расположенный в архиве и выполняем проверки содержимого
def test_csv_file():
    with ZipFile("resources/archive.zip", 'r') as zfile:
        assert "2.csv" in zfile.namelist()
        with zfile.open("2.csv") as csv_file:
            reader = csv.reader(csv_file.read().decode('utf-8').splitlines())
            rows = list(reader)
            # Проверяем, что файл не пустой
            assert len(rows) > 0
            # Проверяем количество строк
            assert len(rows) == 12, "Количество строк не соответствует ожидаемому"
            pprint(rows)



# Читаем xlsx-файл расположенный в архиве и выполняем проверки содержимого.
def test_xlsx():
    with ZipFile("resources/archive.zip", 'r') as zfile_xlsx:
        # Открываем xlsx-файл внутри архива
        with zfile_xlsx.open("1.xlsx") as xlsx_file:
            file_content = io.BytesIO(xlsx_file.read())
            workbook = load_workbook(file_content)
            sheet = workbook.active
            assert sheet['B1'].value == "Январь"
            assert sheet['A12'].value == "=A11+1"


            #в таблице открыть
            for row in sheet.iter_rows(values_only=True):
                print(" | ".join(str(cell) if cell is not None else "" for cell in row))


def test_pdf():
    with ZipFile("resources/archive.zip", 'r') as zfile_pdf:
        # Открываем pdf-файл внутри архива
        with zfile_pdf.open("Python Testing with Pytest (Brian Okken).pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            # Проверяем текст на нужной странице (например, страница 2)
            page_index = 1  # Индекс страницы (начинается с 0)
            page_text = reader.pages[page_index].extract_text()
            # Проверяем, что текст содержит название книги
            assert "Python Testing with pytest" in page_text

            print(page_text)